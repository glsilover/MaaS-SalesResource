# 文件名列表：
import datetime

File_USS = r"/Users/guantianyun/Desktop/20230523205375190687.xlsx"
File_ListOfProduct = r"/Users/guantianyun/Desktop/拆好的表1.xlsx"
File_OutPut = f'/Users/guantianyun/Desktop/汇总账单{datetime.date.today()}.xlsx'

# MaaS 总水位
KEC_CPU_total = 1456
KEC_Mem_total = 2548
Kafka_Mem_total = 2184
RDS_Mem_total = 1152
Redis_Mem_total = 1152
EBS_disk_total = 151000 # GB
KS3_sum = 7 # TB
KS3_total = 620 # TB
KS3_ratio = KS3_sum / KS3_total

# MaaS Price
Price_NAT_IP_Count = 50
Price_NAT_IP_Bandwidth = 100
Price_EIP = 100
Price_MySQL_Mem = 110
Price_MySQL_Disk = 0.4
Price_PG_Mem = 409.38
Price_PG_Disk = 2.67
Price_KEC_CPU = 221
Price_KEC_Mem = 55
Price_Kafka_CPU = 225
Price_Kafka_Mem = 56
Price_EBS = 2.67
Price_Redis = 493.06

# MaaS Discount:
Discount_NAT = 1
Discount_EIP = 1
Discount_MySQL_Double = 1  # 高可用版本
Discount_MySQL_Single = 1  # 单机版
Discount_MySQL_Read = 1  # 只读实例
Discount_MySQL_Free = 1  # 临时实例
Discount_PG_Read = 0.44  # 只读实例
Discount_PG_Double = 0.35  # 高可用版本
Discount_PG_Free = 1  # 临时实例
Discount_KEC = 0.56
Discount_Kafka = 0.57
Discount_EBS = 0.56
Discount_Redis_Double = 0.3
Discount_Redis_Cluster = 0.6






import pandas as pd
import xlrd



# 将原始表格载入"
data = pd.read_excel(File_USS)

# 将计费结束时间、计费开始时间为NULL替换为当前时间
import datetime
now_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
data["计费结束时间"] = data["计费结束时间"].fillna(value = now_time)
data["计费开始时间"] = data["计费开始时间"].fillna(value = now_time)

# 将计费开始/结束时间转换为时间格式
data["计费结束时间"] = pd.to_datetime(data["计费结束时间"])
data["计费开始时间"] = pd.to_datetime(data["计费开始时间"])

# 添加新列"计费时长"
data["计费时长(days)"] = (data["计费结束时间"] - data["计费开始时间"]).dt.days


# 按产品线维度拆分为不同sheet，写入新表"拆好的表.xlsx"
area_list = list(set(data['产品线'])) # print("['托管Hadoop', '弹性IP', '关系型数据库', '容灾部署', '云服务器', '共享带宽', '云数据库PostgreSQL', '网络地址转换NAT', '云硬盘', '云数据库Redis', '负载均衡', '日志服务', '容器引擎']")
writer = pd.ExcelWriter(File_ListOfProduct, engine='xlsxwriter')
data.to_excel(writer, sheet_name="总表", index=False)
for j in area_list:
    df = data[data['产品线'] == j]
    if j == '云服务器':
        df_KEC = df[df['产品类型'] == '通用型N2']
        df_KEC.to_excel(writer, sheet_name='云服务器', index=False)
        df_Kafka = df[df['产品类型'] == 'IO优化型I3']
        df_Kafka.to_excel(writer, sheet_name='托管Kafka', index=False)
    else:
        df.to_excel(writer, sheet_name=j, index=False)
writer.save()  # 一定要加上这句代码，“拆好的表”才会显示出来


# 网络地址转换NAT 处理(非自然月，按30天计算，有误差)
df_NAT = pd.read_excel(File_ListOfProduct, sheet_name='网络地址转换NAT')
df_NAT['带宽(/Mbps)'] = df_NAT['配置详情'].map(lambda x: x.split(': ')[2])
df_NAT['带宽(/Mbps)'] = df_NAT['带宽(/Mbps)'].str.extract('(\d+)', expand=False).map(lambda x: int(x))
df_NAT['IP地址数量'] = df_NAT['配置详情'].map(lambda x: x.split(': ')[3])
df_NAT['IP地址数量'] = df_NAT['IP地址数量'].str.extract('(\d+)', expand=False).map(lambda x: int(x))
df_NAT['原价(元/月)']= (df_NAT['带宽(/Mbps)'] * Price_NAT_IP_Bandwidth + df_NAT['IP地址数量'] * Price_NAT_IP_Count) / 30 * df_NAT["计费时长(days)"]
df_NAT['折扣'] = Discount_NAT
df_NAT['成交价(元/月)'] = df_NAT['原价(元/月)'] * Discount_NAT


# 弹性IP 处理
df_EIP = pd.read_excel(File_ListOfProduct, sheet_name='弹性IP')
df_EIP['带宽(/Mbps)'] = df_EIP['配置详情'].map(lambda x:x.split(': ')[-1])
df_EIP['带宽(/Mbps)'] = df_EIP['带宽(/Mbps)'].str.extract('(\d+)', expand=False).map(lambda x:float(x))
df_EIP['原价(元/月)'] = (df_EIP['带宽(/Mbps)'] * Price_EIP + 50) / 30 * df_EIP["计费时长(days)"]
df_EIP['折扣'] = Discount_EIP
df_EIP['成交价(元/月)'] = df_EIP['原价(元/月)'] * Discount_EIP



# 关系型数据库 处理
df_MySQL = pd.read_excel(File_ListOfProduct, sheet_name='关系型数据库')
df_MySQL['MySQL版本'] = df_MySQL['配置详情'].map(lambda x: x.split('\n')[-6])
df_MySQL['MySQL版本'] = df_MySQL['MySQL版本'].map(lambda x: x.split(' ')[-1])
df_MySQL['内存(/GB)'] = df_MySQL['配置详情'].map(lambda x: x.split('\n')[-5])
df_MySQL['内存(/GB)'] = df_MySQL['内存(/GB)'].str.extract('(\d+)', expand=False).map(lambda x: int(x))
df_MySQL['硬盘(/GB)'] = df_MySQL['配置详情'].map(lambda x: x.split('\n')[-4])
df_MySQL['硬盘(/GB)'] = df_MySQL['硬盘(/GB)'].str.extract('(\d+)', expand=False).map(lambda x: int(x))

df_MySQL_Single = (df_MySQL[df_MySQL['产品类型'] == '单机版']).copy()
df_MySQL_Double = (df_MySQL[df_MySQL['产品类型'] == '高可用版']).copy()
df_MySQL_Read = (df_MySQL[df_MySQL['产品类型'] == '只读RDS']).copy()
df_MySQL_Free = (df_MySQL[df_MySQL['产品类型'] == '临时RDS']).copy()

df_MySQL_Double['原价(元/月)'] = (df_MySQL_Double['内存(/GB)'] * Price_MySQL_Mem + df_MySQL_Double['硬盘(/GB)'] * Price_MySQL_Disk) / 30 * df_MySQL_Double["计费时长(days)"]
df_MySQL_Single['原价(元/月)'] = ((df_MySQL_Single['内存(/GB)'] * Price_MySQL_Mem + df_MySQL_Single['硬盘(/GB)'] * Price_MySQL_Disk) / 30 * df_MySQL_Single["计费时长(days)"]) / 2
df_MySQL_Read['原价(元/月)'] = ((df_MySQL_Read['内存(/GB)'] * Price_MySQL_Mem + df_MySQL_Read['硬盘(/GB)'] * Price_MySQL_Disk) / 30 * df_MySQL_Read["计费时长(days)"]) / 2
df_MySQL_Free['原价(元/月)'] = 0
df_MySQL_Double['折扣'] = Discount_MySQL_Double
df_MySQL_Single['折扣'] = Discount_MySQL_Single
df_MySQL_Read['折扣'] = Discount_MySQL_Read
df_MySQL_Free['折扣'] = Discount_MySQL_Free
df_MySQL = pd.concat([df_MySQL_Double, df_MySQL_Single, df_MySQL_Read, df_MySQL_Free], ignore_index=True)
df_MySQL['成交价(元/月)'] = df_MySQL['原价(元/月)'] * df_MySQL['折扣']



# 云数据库PostgreSQL 处理
df_PG = pd.read_excel(File_ListOfProduct, sheet_name='云数据库PostgreSQL')
df_PG['PG版本'] = df_PG['配置详情'].map(lambda x:x.split('\n')[-6])
df_PG['PG版本'] = df_PG['PG版本'].map(lambda x:x.split(' ')[-1])
df_PG['内存(/GB)'] = df_PG['配置详情'].map(lambda x:x.split('\n')[-5])
df_PG['内存(/GB)'] = df_PG['内存(/GB)'].str.extract('(\d+)', expand=False).map(lambda x:int(x))
df_PG['硬盘(/GB)'] = df_PG['配置详情'].map(lambda x:x.split('\n')[-4])
df_PG['硬盘(/GB)'] = df_PG['硬盘(/GB)'].str.extract('(\d+)', expand=False).map(lambda x:int(x))

df_PG_Double = (df_PG[df_PG['产品类型'] == '高可用版']).copy()
df_PG_Read = (df_PG[df_PG['产品类型'] == '只读实例']).copy()
df_PG_Free = (df_PG[df_PG['产品类型'] == '临时版']).copy()

df_PG_Double['原价(元/月)'] = (df_PG_Double['内存(/GB)'] * Price_PG_Mem + df_PG_Double['硬盘(/GB)'] * Price_PG_Disk) / 30 * df_PG_Double["计费时长(days)"]
df_PG_Read['原价(元/月)'] = ((df_PG_Read['内存(/GB)'] * Price_PG_Mem + df_PG_Read['硬盘(/GB)'] * Price_PG_Disk) / 30 * df_PG_Read["计费时长(days)"]) / 2
df_PG_Free['原价(元/月)'] = 0
df_PG_Double['折扣'] = Discount_PG_Double
df_PG_Read['折扣'] = Discount_PG_Read
df_PG_Free['折扣'] = Discount_PG_Free
df_PG = pd.concat([df_PG_Double, df_PG_Read, df_PG_Free], ignore_index=True)
df_PG['成交价(元/月)'] = df_PG['原价(元/月)'] * df_PG['折扣']


# 云服务器 处理
df_KEC = pd.read_excel(File_ListOfProduct, sheet_name='云服务器')
df_KEC['配置详情'] = df_KEC['配置详情'].map(lambda x:x.split('CPU: ')[-1])
df_KEC['CPU(/C)'] = df_KEC['配置详情'].map(lambda x:x.split('核')[0]).map(lambda x:int(x))
df_KEC['配置详情'] = df_KEC['配置详情'].map(lambda x:x.split('内存: ')[-1])
df_KEC['内存(/GB)'] = df_KEC['配置详情'].map(lambda x:x.split('GB')[0]).map(lambda x:int(x))

df_KEC['原价(元/月)'] = (df_KEC['CPU(/C)'] * Price_KEC_CPU + df_KEC['内存(/GB)'] * Price_KEC_Mem) / 30 * df_KEC["计费时长(days)"]
df_KEC['折扣'] = Discount_KEC
df_KEC['成交价(元/月)'] = df_KEC['原价(元/月)'] * Discount_KEC


# 托管Kafka 处理
df_Kafka = pd.read_excel(File_ListOfProduct, sheet_name='托管Kafka')
df_Kafka_sum = df_Kafka.copy()
df_Kafka['配置详情'] = df_Kafka['配置详情'].map(lambda x:x.split('CPU: ')[-1])
df_Kafka['CPU(/C)'] = df_Kafka['配置详情'].map(lambda x:x.split('核')[0]).map(lambda x:int(x))
df_Kafka['配置详情'] = df_Kafka['配置详情'].map(lambda x:x.split('内存: ')[-1])
df_Kafka['内存(/GB)'] = df_Kafka['配置详情'].map(lambda x:x.split('GB')[0]).map(lambda x:int(x))

df_Kafka['原价(元/月)'] = (df_Kafka['CPU(/C)'] * Price_Kafka_CPU + df_Kafka['内存(/GB)'] * Price_Kafka_Mem + 40 * Price_EBS) / 30 * df_Kafka["计费时长(days)"]
df_Kafka['折扣'] = Discount_Kafka
df_Kafka['成交价(元/月)'] = ((df_Kafka['CPU(/C)'] * Price_Kafka_CPU + df_Kafka['内存(/GB)'] * Price_Kafka_Mem) * Discount_Kafka + (40 * Price_EBS) * Discount_EBS) / 30 * df_Kafka["计费时长(days)"]


# 云硬盘 处理
df_EBS = pd.read_excel(File_ListOfProduct, sheet_name='云硬盘')
df_EBS['硬盘(/GB)'] = df_EBS['配置详情'].map(lambda x:x.split('容量: ')[-1])
df_EBS['硬盘(/GB)'] = df_EBS['硬盘(/GB)'].str.extract('(\d+)', expand=False).map(lambda x:int(x))

df_EBS['原价(元/月)'] = (df_EBS['硬盘(/GB)'] * Price_EBS + 50) / 30 * df_EBS["计费时长(days)"]
df_EBS['折扣'] = Discount_EBS
df_EBS['成交价(元/月)'] = df_EBS['原价(元/月)'] * Discount_EBS


# 云数据库Redis 处理(只读实例无法计费)
df_Redis = pd.read_excel(File_ListOfProduct, sheet_name='云数据库Redis')
df_Redis['内存(/GB)'] = df_Redis['配置详情'].map(lambda x:x.split('连接数:')[0])
df_Redis['内存(/GB)'] = df_Redis['内存(/GB)'].map(lambda x:x.split('内存容量:')[-1])
df_Redis['内存(/GB)'] = df_Redis['内存(/GB)'].str.extract('(\d+)', expand=False).map(lambda x:int(x))

df_Redis_Double = (df_Redis[df_Redis['产品类型'] == '主从']).copy()
df_Redis_Cluster = (df_Redis[df_Redis['产品类型'] == '自定义集群']).copy()

df_Redis_Double['原价(元/月)'] = (df_Redis_Double['内存(/GB)'] * Price_Redis) / 30 * df_Redis_Double["计费时长(days)"]
df_Redis_Cluster['原价(元/月)'] = ((df_Redis_Cluster['内存(/GB)'] * Price_Redis) / 30 * df_Redis_Cluster["计费时长(days)"]) / 2
df_Redis_Double['折扣'] = Discount_Redis_Double
df_Redis_Cluster['折扣'] = Discount_Redis_Cluster
df_Redis = pd.concat([df_Redis_Double, df_Redis_Cluster], ignore_index=True)

df_Redis['成交价(元/月)'] = df_Redis['原价(元/月)'] * df_Redis['折扣']

df_sum = pd.concat([df_NAT, df_EIP, df_KEC, df_Kafka, df_Redis, df_EBS, df_PG, df_MySQL], ignore_index=True)

for i in ('带宽(/Mbps)', 'IP地址数量', 'CPU(/C)', '内存(/GB)', '硬盘(/GB)', 'PG版本', 'MySQL版本'):
    del df_sum[i]


# 定义写操作
writer_finally = pd.ExcelWriter(File_OutPut, engine='xlsxwriter')
# 分别将表df1、df2、df
# 3写入Excel中的sheet1、sheet2、sheet3
# 命名为表1、表2、表3
df_KEC.to_excel(writer_finally, sheet_name='KEC', index=False)
df_EIP.to_excel(writer_finally, sheet_name='EIP', index=False)
df_NAT.to_excel(writer_finally, sheet_name='NAT', index=False)
df_EBS.to_excel(writer_finally, sheet_name='EBS', index=False)
df_PG.to_excel(writer_finally, sheet_name='PG', index=False)
df_MySQL.to_excel(writer_finally, sheet_name='MySQL', index=False)
df_Redis.to_excel(writer_finally, sheet_name='Redis', index=False)
df_Kafka.to_excel(writer_finally, sheet_name='Kafka', index=False)
df_sum.to_excel(writer_finally, sheet_name='sum', index=False)
# 保存读写的内容
writer_finally.save()

# KEC 资源利用率计算
df_KEC_use = df_KEC[df_KEC['产品类型'] == '通用型N2']
df_KEC_use = df_KEC_use[df_KEC_use['服务状态'] == '已开通']
KEC_CPU_sum = df_KEC_use['CPU(/C)'].sum()
KEC_Mem_sum = df_KEC_use['内存(/GB)'].sum()
KEC_CPU_ratio = KEC_CPU_sum / KEC_CPU_total * 100
KEC_Mem_ratio = KEC_Mem_sum / KEC_Mem_total * 100


# RDS 资源利用率统计
def RDS_userage(df, RDS_banben, RDS_leixing):
    df1 = df[df['服务状态'] == '已开通']
    df_double = df1[df1['产品类型'] == RDS_banben]
    df_double_sum = df_double[RDS_leixing].sum()
    df_single = df1[df1['产品类型'] != RDS_banben]
    df_single_sum = df_single[RDS_leixing].sum()
    return df_single_sum + df_double_sum * 2

RDS_sum = RDS_userage(df_MySQL, '高可用版', '内存(/GB)') + RDS_userage(df_PG, '高可用版', '内存(/GB)')
RDS_ratio = RDS_sum / RDS_Mem_total * 100


# Redis 资源利用率统计
Redis_sum = RDS_userage(df_Redis, '主从', '内存(/GB)')
Redis_ratio = Redis_sum / Redis_Mem_total * 100


# 云硬盘统计
EBS_sum = RDS_userage(df_EBS, '1', '硬盘(/GB)')
EBS_ratio = EBS_sum / EBS_disk_total * 100


# Kafka资源统计
df_Kafka_use = df_Kafka[df_Kafka['服务状态'] == '已开通']
Kafka_Mem_sum = df_Kafka_use['内存(/GB)'].sum()
Kafka_Mem_ratio = Kafka_Mem_sum / Kafka_Mem_total * 100






# 将资源利用率数据导入一个新的sheet
from openpyxl import load_workbook

# 加载excel，注意路径要与脚本一致
wb = load_workbook(File_OutPut)
wb.create_sheet(title='ratio',index=0)
# 激活excel表
sheet = wb.active

# 向excel中写入表头
sheet['a1'] = '云主机-CPU'
sheet['b1'] = '云主机-内存'
sheet['c1'] = 'RDS'
sheet['d1'] = 'Redis'
sheet['e1'] = 'Kafka'
sheet['f1'] = 'EBS'
sheet['g1'] = '对象存储'
# 向excel中写入对应的value
sheet.cell(row=2, column=1).value = round(KEC_CPU_ratio, 0)
sheet.cell(row=2, column=2).value = round(KEC_Mem_ratio, 0)
sheet.cell(row=2, column=3).value = round(RDS_ratio, 0)
sheet.cell(row=2, column=4).value = round(Redis_ratio, 0)
sheet.cell(row=2, column=5).value = round(Kafka_Mem_ratio, 0)
sheet.cell(row=2, column=6).value = round(EBS_ratio, 0)
sheet.cell(row=2, column=7).value = round(KS3_ratio, 0)


wb.create_sheet(title='usage',index=0)
# 激活excel表
sheet_use = wb.active

sheet_use['a1'] = '品类'
sheet_use['b1'] = '用量/总量'

sheet_use.cell(row=2, column=1).value = '云主机-CPU'
sheet_use.cell(row=3, column=1).value = '云主机-内存'
sheet_use.cell(row=4, column=1).value = 'RDS'
sheet_use.cell(row=5, column=1).value = 'Redis'
sheet_use.cell(row=6, column=1).value = 'Kafka'
sheet_use.cell(row=7, column=1).value = 'EBS'
sheet_use.cell(row=8, column=1).value = '对象存储'

def sheet_out(r, c, a, b):
    a = int(a)
    b = int(b)
    sheet_use.cell(row=r, column=c).value = b - a

sheet_out(2, 2, KEC_CPU_sum, KEC_CPU_total)
sheet_out(3, 2, KEC_Mem_sum, KEC_Mem_total)
sheet_out(4, 2, RDS_sum, RDS_Mem_total)
sheet_out(5, 2, Redis_sum, Redis_Mem_total)
sheet_out(6, 2, Kafka_Mem_sum, Kafka_Mem_total)
sheet_out(7, 2, EBS_sum / 1000, EBS_disk_total / 1000)
sheet_out(8, 2, KS3_sum, KS3_total)

wb.save(File_OutPut)
print('数据写入成功！')